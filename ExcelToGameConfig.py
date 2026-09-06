import ast
import hashlib
import json
import os
import re
import struct
import subprocess
import sys
import traceback
from datetime import datetime
from time import perf_counter
from typing import Any, Dict, List

import openpyxl

TYPE_INT = "int"
TYPE_STRING = "string"
TYPE_FLOAT = "float"
TYPE_FLOAT_ARR = "floatArr"
TYPE_INT_ARR = "intArr"
TYPE_STRING_ARR = "stringArr"
TYPE_FLOAT_ARR2 = "floatArr2"
TYPE_INT_ARR2 = "intArr2"
TYPE_STRING_ARR2 = "stringArr2"
SUPPORTED_TYPES = {
    TYPE_INT,
    TYPE_STRING,
    TYPE_FLOAT,
    TYPE_FLOAT_ARR,
    TYPE_INT_ARR,
    TYPE_STRING_ARR,
    TYPE_FLOAT_ARR2,
    TYPE_INT_ARR2,
    TYPE_STRING_ARR2,
}
BACKUP_NAME_MARKER = "备份"
LARGE_SHEET_ROW_THRESHOLD = 100_000
BLANK_ROW_STOP_THRESHOLD = 256


def _default_base_dir() -> str:
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def _default_excel_dir(base_dir: str) -> str:
    return base_dir

def _default_output_dir(base_dir: str) -> str:
    return os.path.join(base_dir, "Config")


def _safe_mkdir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _show_result_dialog(title: str, message: str, success: bool, auto_close_ms: int = 0) -> None:
    """Show an export result dialog with Tk and macOS native fallbacks."""
    try:
        import tkinter as tk
        from tkinter import ttk

        root = tk.Tk()
        root.title(title)
        root.attributes("-topmost", True)
        root.resizable(False, False)

        frame = ttk.Frame(root, padding=20)
        frame.grid(row=0, column=0, sticky="nsew")

        heading = "导表成功" if success else "导表失败"
        heading_label = ttk.Label(frame, text=heading, font=("Microsoft YaHei UI", 14, "bold"))
        heading_label.grid(row=0, column=0, sticky="w")

        message_label = ttk.Label(
            frame,
            text=message,
            font=("Microsoft YaHei UI", 10),
            justify="left",
            wraplength=560,
        )
        message_label.grid(row=1, column=0, sticky="w", pady=(12, 0))

        if success:
            timeout_label = ttk.Label(frame, text="窗口将在 3 秒后自动关闭")
            timeout_label.grid(row=2, column=0, sticky="w", pady=(14, 0))
        else:
            close_button = ttk.Button(frame, text="确定", command=root.destroy)
            close_button.grid(row=2, column=0, sticky="e", pady=(18, 0))
            close_button.focus_set()

        root.update_idletasks()
        width = root.winfo_reqwidth()
        height = root.winfo_reqheight()
        x = max(0, (root.winfo_screenwidth() - width) // 2)
        y = max(0, (root.winfo_screenheight() - height) // 2)
        root.geometry(f"+{x}+{y}")
        root.lift()

        if auto_close_ms > 0:
            root.after(auto_close_ms, root.destroy)
        root.mainloop()
    except Exception as dialog_error:
        # PyInstaller windowed apps may not include a working Tcl/Tk runtime.
        # Use macOS's built-in dialog so double-click failures remain visible.
        try:
            def _apple_script_text(value: str) -> str:
                return value.replace("\\", "\\\\").replace('"', '\\"').replace("\n", "\\n")

            script = (
                f'display dialog "{_apple_script_text(message)}" '
                f'with title "{_apple_script_text(title)}" '
                f'buttons {{"确定"}} default button "确定"'
            )
            if success and auto_close_ms > 0:
                script += f" giving up after {max(1, auto_close_ms // 1000)}"
            subprocess.run(["/usr/bin/osascript", "-e", script], check=False)
        except Exception as native_dialog_error:
            print(f"无法显示结果弹窗: {dialog_error}; macOS 弹窗失败: {native_dialog_error}", file=sys.stderr)


def _format_export_error(exc: BaseException) -> str:
    chain: List[BaseException] = []
    current: BaseException | None = exc
    while current is not None and current not in chain:
        chain.append(current)
        current = current.__cause__ or current.__context__

    root_cause = chain[-1]
    detail = str(exc).strip() or repr(exc)

    if isinstance(root_cause, PermissionError):
        target = getattr(root_cause, "filename", None)
        reason = "没有权限访问目标文件，文件也可能正被 Excel 或其他程序占用。"
        if target:
            reason += f"\n目标文件：{target}"
    elif isinstance(root_cause, FileNotFoundError):
        reason = "没有找到需要的文件或目录。"
    elif isinstance(root_cause, OSError) and getattr(root_cause, "errno", None) == 28:
        reason = "磁盘空间不足，无法写入导表结果。"
    else:
        reason = detail

    if reason != detail:
        reason += f"\n\n详细信息：{detail}"
    return f"{reason}\n\n错误类型：{root_cause.__class__.__name__}"


def _short_value(value: Any, limit: int = 120) -> str:
    text = repr(value)
    return text if len(text) <= limit else text[: limit - 3] + "..."


def _is_excel_file(name: str) -> bool:
    lower = name.lower()
    return not lower.startswith("~$") and (lower.endswith(".xlsx") or lower.endswith(".xlsm"))


def _is_backup_name(name: str) -> bool:
    return BACKUP_NAME_MARKER in name


def _parse_json_text(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value
        if text == "":
            return ""
        text = text.strip()
        if not text:
            return value
        if (text.startswith("[") and text.endswith("]")) or (text.startswith("{") and text.endswith("}")):
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                try:
                    return ast.literal_eval(text)
                except (ValueError, SyntaxError):
                    return value
        return value
    return value


def _parse_array_text(value: Any, typ: str) -> Any:
    """Parse JSON arrays and the comma-delimited notation common in Excel."""
    parsed = _parse_json_text(value)
    if not isinstance(parsed, str):
        return parsed

    text = parsed.strip()
    if text == "":
        return []

    if typ in (TYPE_INT_ARR2, TYPE_FLOAT_ARR2, TYPE_STRING_ARR2):
        rows = re.split(r"[;；\n]+", text)
        return [re.split(r"[,，]+", row.strip()) for row in rows if row.strip()]

    if typ in (TYPE_INT_ARR, TYPE_FLOAT_ARR, TYPE_STRING_ARR):
        return [item.strip() for item in re.split(r"[,，;；\n]+", text) if item.strip()]

    return parsed


def _cast_scalar(value: Any, typ: str) -> Any:
    if value is None:
        return None
    if typ == TYPE_STRING:
        if isinstance(value, str):
            if value.strip() == '""':
                return ""
            return value
        return str(value)
    if typ == TYPE_INT:
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            text = value.strip()
            if text == "":
                return None
            return int(float(text))
        return int(value)
    if typ == TYPE_FLOAT:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            text = value.strip()
            if text == "":
                return None
            return float(text)
        return float(value)
    return value


def _ensure_array_value(value: Any, typ: str) -> List[Any]:
    if value is None:
        return []
    if isinstance(value, (list, tuple)):
        return list(value)
    return [_cast_scalar(value, typ)]


def _ensure_array2_value(value: Any) -> List[List[Any]]:
    if value is None:
        return []
    if not isinstance(value, (list, tuple)):
        raise ValueError(f"Expected 2D array value, got {type(value).__name__}: {value!r}")
    rows = list(value)
    for row in rows:
        if not isinstance(row, (list, tuple)):
            raise ValueError(f"Expected each 2D array row to be a list, got {type(row).__name__}: {row!r}")
    return [list(row) for row in rows]


def _cast_by_type(value: Any, typ: str) -> Any:
    if value is None:
        return None
    if typ in (TYPE_STRING, TYPE_INT, TYPE_FLOAT):
        return _cast_scalar(value, typ)
    parsed = _parse_array_text(value, typ)
    if typ == TYPE_INT_ARR:
        return [int(float(x)) for x in _ensure_array_value(parsed, TYPE_INT)]
    if typ == TYPE_FLOAT_ARR:
        return [float(x) for x in _ensure_array_value(parsed, TYPE_FLOAT)]
    if typ == TYPE_STRING_ARR:
        return [str(x) for x in _ensure_array_value(parsed, TYPE_STRING)]
    if typ == TYPE_INT_ARR2:
        return [[int(float(x)) for x in row] for row in _ensure_array2_value(parsed)]
    if typ == TYPE_FLOAT_ARR2:
        return [[float(x) for x in row] for row in _ensure_array2_value(parsed)]
    if typ == TYPE_STRING_ARR2:
        return [[str(x) for x in row] for row in _ensure_array2_value(parsed)]
    return parsed


def _sheet_to_table(ws) -> Dict[str, Any]:
    start_time = perf_counter()
    cols: List[int] = []
    field_names: List[str] = []
    field_types: List[str] = []
    max_row = ws.max_row or 0
    max_column = ws.max_column or 0

    # Row 1: comment only, ignored by exporter.
    # Row 2: field names.
    # Row 3: field types.
    header_rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    name_row = header_rows[1] if len(header_rows) >= 2 else ()
    type_row = header_rows[2] if len(header_rows) >= 3 else ()
    header_width = max(len(name_row), len(type_row), max_column)

    for idx in range(header_width):
        key = name_row[idx] if idx < len(name_row) else None
        typ = type_row[idx] if idx < len(type_row) else None
        if key is None or str(key).strip() == "":
            continue
        if typ is None:
            continue
        typ_text = str(typ).strip()
        if typ_text == "" or typ_text.lower() == "null":
            # null type means comment-only/export-ignore column.
            continue
        if typ_text not in SUPPORTED_TYPES:
            raise ValueError(f"Unsupported type '{typ_text}' in sheet '{ws.title}', column '{key}'")
        cols.append(idx + 1)
        field_names.append(str(key).strip())
        field_types.append(typ_text)

    rows: List[Dict[str, Any]] = []
    scanned_rows = 0
    blank_run = 0
    early_stopped = False
    large_sheet = max_row >= LARGE_SHEET_ROW_THRESHOLD

    for row_values in ws.iter_rows(min_row=4, values_only=True):
        scanned_rows += 1
        row_obj: Dict[str, Any] = {}
        empty = True
        for idx, key in enumerate(field_names):
            col_index = cols[idx] - 1
            raw = row_values[col_index] if col_index < len(row_values) else None
            raw_blank = raw is None or (isinstance(raw, str) and raw.strip() == "")
            if not raw_blank:
                empty = False
            if raw_blank:
                continue
            try:
                row_obj[key] = _cast_by_type(raw, field_types[idx])
            except (TypeError, ValueError, OverflowError) as exc:
                excel_row = scanned_rows + 3
                raise ValueError(
                    f"数据转换失败：工作表 '{ws.title}'，第 {excel_row} 行，"
                    f"字段 '{key}'，声明类型 '{field_types[idx]}'，原始值 {_short_value(raw)}"
                ) from exc
        if not empty:
            rows.append(row_obj)
            blank_run = 0
        else:
            blank_run += 1
            if large_sheet and blank_run >= BLANK_ROW_STOP_THRESHOLD:
                early_stopped = True
                break

    elapsed = perf_counter() - start_time
    if early_stopped:
        print(
            f"Warning: sheet '{ws.title}' reports max_row={max_row}; "
            f"stopped after {blank_run} consecutive blank rows to avoid used-range bloat."
        )
    print(
        f"Sheet: {ws.title} | max_row={max_row} | scanned_rows={scanned_rows} | "
        f"export_rows={len(rows)} | fields={len(field_names)} | elapsed={elapsed:.3f}s"
    )

    return {
        "name": ws.title,
        "field_names": field_names,
        "field_types": field_types,
        "rows": rows,
    }


def build_tables(input_dir: str) -> List[Dict[str, Any]]:
    tables: List[Dict[str, Any]] = []
    existing_sheet_names = set()
    excel_files = [
        os.path.join(input_dir, n)
        for n in os.listdir(input_dir)
        if _is_excel_file(n) and not _is_backup_name(n)
    ]
    excel_files.sort()
    if not excel_files:
        raise FileNotFoundError(f"No Excel files found in: {input_dir}")

    total_start = perf_counter()
    for excel_path in excel_files:
        excel_start = perf_counter()
        print(f"Loading workbook: {os.path.basename(excel_path)}")
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
            try:
                for sheet_name in wb.sheetnames:
                    if _is_backup_name(sheet_name):
                        # Backup sheets are intentionally excluded from game_config.bin.
                        continue
                    if sheet_name.lower().endswith("_map"):
                        # Mapping tables are not exported into game_config.bin.
                        continue
                    if sheet_name in existing_sheet_names:
                        raise ValueError(f"多个 Excel 中存在同名工作表：'{sheet_name}'")
                    existing_sheet_names.add(sheet_name)
                    tables.append(_sheet_to_table(wb[sheet_name]))
            finally:
                wb.close()
        except Exception as exc:
            raise RuntimeError(f"处理 Excel 文件 '{os.path.basename(excel_path)}' 失败：{exc}") from exc
        print(f"Workbook done: {os.path.basename(excel_path)} | elapsed={perf_counter() - excel_start:.3f}s")
    print(f"Build tables done: tables={len(tables)} | elapsed={perf_counter() - total_start:.3f}s")
    return tables


def _write_int32_be(f, value: int) -> None:
    f.write(struct.pack(">i", int(value)))


def _write_uint8(f, value: int) -> None:
    f.write(struct.pack(">B", int(value)))


def _write_float_be(f, value: float) -> None:
    f.write(struct.pack(">f", float(value)))


def _write_string(f, value: str) -> None:
    text = "" if value is None else str(value)
    data = text.encode("utf-8")
    _write_int32_be(f, len(data))
    f.write(data)


def _write_int_array(f, values: List[Any]) -> None:
    _write_int32_be(f, len(values))
    for value in values:
        _write_int32_be(f, int(value))


def _write_float_array(f, values: List[Any]) -> None:
    _write_int32_be(f, len(values))
    for value in values:
        _write_float_be(f, float(value))


def _write_string_array(f, values: List[Any]) -> None:
    _write_int32_be(f, len(values))
    for value in values:
        _write_string(f, str(value))


def _write_int_array2(f, values: List[List[Any]]) -> None:
    _write_int32_be(f, len(values))
    for row in values:
        _write_int_array(f, row)


def _write_float_array2(f, values: List[List[Any]]) -> None:
    _write_int32_be(f, len(values))
    for row in values:
        _write_float_array(f, row)


def _write_string_array2(f, values: List[List[Any]]) -> None:
    _write_int32_be(f, len(values))
    for row in values:
        _write_string_array(f, row)


def _write_value(f, typ: str, value: Any) -> None:
    if typ == TYPE_INT:
        _write_int32_be(f, 0 if value is None else int(value))
        return
    if typ == TYPE_FLOAT:
        _write_float_be(f, 0.0 if value is None else float(value))
        return
    if typ == TYPE_STRING:
        _write_string(f, "" if value is None else str(value))
        return
    if typ == TYPE_INT_ARR:
        _write_int_array(f, [] if value is None else value)
        return
    if typ == TYPE_FLOAT_ARR:
        _write_float_array(f, [] if value is None else value)
        return
    if typ == TYPE_STRING_ARR:
        _write_string_array(f, [] if value is None else value)
        return
    if typ == TYPE_INT_ARR2:
        _write_int_array2(f, [] if value is None else value)
        return
    if typ == TYPE_FLOAT_ARR2:
        _write_float_array2(f, [] if value is None else value)
        return
    if typ == TYPE_STRING_ARR2:
        _write_string_array2(f, [] if value is None else value)
        return
    raise ValueError(f"Unsupported field type while writing bin: {typ}")


def write_game_config_bin(output_dir: str, tables: List[Dict[str, Any]]) -> str:
    _safe_mkdir(output_dir)
    path = os.path.join(output_dir, "game_config.bin")
    with open(path, "wb") as f:
        _write_int32_be(f, len(tables))
        for table in tables:
            field_names = table["field_names"]
            field_types = table["field_types"]
            rows = table["rows"]

            _write_string(f, table["name"])
            _write_int32_be(f, len(rows))
            _write_int32_be(f, len(field_names))

            for i in range(len(field_names)):
                _write_string(f, field_names[i])
                _write_string(f, field_types[i])

            for row in rows:
                for i in range(len(field_names)):
                    field_name = field_names[i]
                    field_type = field_types[i]
                    _write_value(f, field_type, row.get(field_name))
    return path


def write_game_config_json(output_dir: str, tables: List[Dict[str, Any]]) -> str:
    _safe_mkdir(output_dir)
    path = os.path.join(output_dir, "game_config.json")
    with open(path, "w", encoding="utf-8") as f:
        json.dump({"tables": tables}, f, ensure_ascii=False, indent=4)
    return path


def write_version_file(output_dir: str, version: str, config_path: str) -> str:
    with open(config_path, "rb") as f:
        raw = f.read()
    sha256 = hashlib.sha256(raw).hexdigest()
    payload = {"version": version, "hash": sha256}
    version_path = os.path.join(output_dir, "config_version.json")
    with open(version_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=4)
    return version_path


def _resolve_version() -> str:
    if len(sys.argv) >= 2 and sys.argv[1].strip():
        return sys.argv[1].strip()
    return datetime.now().strftime("%Y.%m.%d.%H%M%S")


def main() -> int:
    base_dir = _default_base_dir()
    input_dir = _default_excel_dir(base_dir)
    output_dir = _default_output_dir(base_dir)
    version = _resolve_version()

    try:
        tables = build_tables(input_dir)
        config_path = write_game_config_bin(output_dir, tables)
        json_path = write_game_config_json(output_dir, tables)
        version_path = write_version_file(output_dir, version, config_path)

        print("Build game_config.bin success")
        print(f"InputDir: {input_dir}")
        print(f"OutputDir: {output_dir}")
        print(f"Version: {version}")
        print(f"Config: {config_path}")
        print(f"Json: {json_path}")
        print(f"VersionFile: {version_path}")
        print(f"TableCount: {len(tables)}")

        success_message = (
            f"共导出 {len(tables)} 个配置表\n"
            f"输出目录：{output_dir}\n"
            f"版本：{version}"
        )
        _show_result_dialog("导表成功", success_message, success=True, auto_close_ms=3000)
        return 0
    except Exception as exc:
        traceback.print_exc()
        error_message = _format_export_error(exc)
        _show_result_dialog("导表失败", error_message, success=False)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
