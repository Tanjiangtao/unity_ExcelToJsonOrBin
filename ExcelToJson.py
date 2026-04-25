import ast
import json
import os
import sys
from typing import Any, Dict, List, Optional, Union

import openpyxl


TYPE_INT = "int"
TYPE_STRING = "string"
TYPE_FLOAT = "float"
TYPE_FLOAT_ARR = "floatArr"
TYPE_INT_ARR = "intArr"
TYPE_STRING_ARR = "stringArr"
TYPE_FLOAT_ARR2 = "floatArr2"
TYPE_INT_ARR2 = "intArr2"
TYPE_STRING_ARR2 = "stringArr2"


def _parse_json_text(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        text = value
        # keep empty string as "" (do not coerce to None)
        if text == "":
            return ""
        text = text.strip()
        if not text:
            return value
        if (text.startswith("[") and text.endswith("]")) or (text.startswith("{") and text.endswith("}")):
            try:
                return json.loads(text)
            except json.JSONDecodeError:
                # allow python literal style too
                try:
                    return ast.literal_eval(text)
                except (ValueError, SyntaxError):
                    return value
        return value
    return value


def _cast_scalar(value: Any, typ: str) -> Any:
    if value is None:
        return None

    if typ == TYPE_STRING:
        if value is None:
            return None
        if isinstance(value, str):
            # Special token to preserve empty string through Excel
            if value.strip() == '""':
                return ""
            return value
        return str(value)

    if typ == TYPE_INT:
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            return int(value)
        if isinstance(value, str):
            if value == "":
                return None
            text = value.strip()
            if text == "":
                return None
            return int(float(text))
        return int(value)

    if typ == TYPE_FLOAT:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str):
            if value == "":
                return None
            text = value.strip()
            if text == "":
                return None
            return float(text)
        return float(value)

    return value


def _cast_by_type(value: Any, typ: str) -> Any:
    if value is None:
        return None

    if typ in (TYPE_STRING, TYPE_INT, TYPE_FLOAT):
        return _cast_scalar(value, typ)

    parsed = _parse_json_text(value)

    if typ == TYPE_INT_ARR:
        if parsed is None:
            return []
        return [int(float(x)) for x in list(parsed)]

    if typ == TYPE_FLOAT_ARR:
        if parsed is None:
            return []
        return [float(x) for x in list(parsed)]

    if typ == TYPE_STRING_ARR:
        if parsed is None:
            return []
        return [str(x) for x in list(parsed)]

    if typ == TYPE_INT_ARR2:
        if parsed is None:
            return []
        return [[int(float(x)) for x in row] for row in list(parsed)]

    if typ == TYPE_FLOAT_ARR2:
        if parsed is None:
            return []
        return [[float(x) for x in row] for row in list(parsed)]

    if typ == TYPE_STRING_ARR2:
        if parsed is None:
            return []
        return [[str(x) for x in row] for row in list(parsed)]

    # unknown type: return as-is
    return parsed


def _set_nested(target: Dict[str, Any], parts: List[str], value: Any) -> None:
    cur: Any = target
    for key in parts[:-1]:
        nxt = cur.get(key)
        if not isinstance(nxt, dict):
            nxt = {}
            cur[key] = nxt
        cur = nxt
    cur[parts[-1]] = value


def excel_to_json(
    excel_file_path: str,
    json_file_path: Optional[str] = None,
    sheet_name: Union[None, int, str, List[str]] = None,
    sep: str = "_",
) -> Dict[str, List[Dict[str, Any]]]:
    """Excel -> JSON using 3 header rows.

    Row 1: comment (ignored)
    Row 2: key
    Row 3: type
    Row 4+: data

    Note: We DO NOT auto-expand keys by splitting on "_" because your source keys
    already contain many underscores (e.g. attack_cd, name_icon). Keeping keys as-is
    avoids wrong nesting.
    """

    if not os.path.exists(excel_file_path):
        raise FileNotFoundError(f"Error: Excel file not found -> {excel_file_path}")

    wb = openpyxl.load_workbook(excel_file_path, data_only=True)

    if sheet_name is None:
        sheet_names = wb.sheetnames
    elif isinstance(sheet_name, int):
        sheet_names = [wb.sheetnames[sheet_name]]
    elif isinstance(sheet_name, str):
        sheet_names = [sheet_name]
    else:
        sheet_names = list(sheet_name)

    output: Dict[str, List[Dict[str, Any]]] = {}

    for name in sheet_names:
        ws = wb[name]

        # read header rows
        cols: List[int] = []
        keys: List[str] = []
        types: List[str] = []

        max_col = ws.max_column
        for c in range(1, max_col + 1):
            key = ws.cell(row=2, column=c).value
            typ = ws.cell(row=3, column=c).value

            if key is None or str(key).strip() == "":
                continue

            # If type is "null" (or blank), treat this column as comment-only
            # and do not export it to JSON.
            if typ is None:
                continue
            typ_text = str(typ).strip()
            if typ_text == "" or typ_text.lower() == "null":
                continue

            cols.append(c)
            keys.append(str(key).strip())
            types.append(typ_text)

        records: List[Dict[str, Any]] = []

        for r in range(4, ws.max_row + 1):
            row_obj: Dict[str, Any] = {}
            empty = True

            for idx, key in enumerate(keys):
                c = cols[idx]
                raw = ws.cell(row=r, column=c).value

                raw_blank = raw is None or (isinstance(raw, str) and raw.strip() == "")
                if not raw_blank:
                    empty = False

                # If the cell is blank, skip the field entirely so we don't
                # introduce new keys that did not exist in the original JSON.
                if raw_blank:
                    continue

                typ = types[idx]
                value = _cast_by_type(raw, typ)
                row_obj[key] = value

            if not empty:
                records.append(row_obj)

        output[name] = records

    if json_file_path is None:
        json_file_path = os.path.splitext(excel_file_path)[0] + ".json"

    with open(json_file_path, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=4)

    print("Excel to JSON done")
    print(f"Source Excel: {excel_file_path}")
    print(f"Output JSON: {json_file_path}")

    return output


def _default_input_dir() -> str:
    # When packaged by PyInstaller, sys.argv[0] points to the exe path.
    # When running as .py, it's the script path.
    return os.path.dirname(os.path.abspath(sys.argv[0]))


def _safe_mkdir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def _is_excel_file(name: str) -> bool:
    lower = name.lower()
    if lower.startswith("~$"):
        return False
    return lower.endswith(".xlsx") or lower.endswith(".xlsm")


def _notify_windows(title: str, message: str) -> None:
    try:
        import ctypes

        MB_OK = 0x0
        MB_ICONINFORMATION = 0x40
        MB_ICONWARNING = 0x30
        flags = MB_OK | (MB_ICONWARNING if "failed" in title.lower() or "失败" in title else MB_ICONINFORMATION)
        ctypes.windll.user32.MessageBoxW(None, message, title, flags)
    except Exception:
        # If running in a non-Windows environment or messagebox fails, ignore.
        return


def convert_all_excels_in_dir(input_dir: str, output_dir: str) -> int:
    """Convert all Excel files in input_dir to JSON in output_dir.

    Returns process exit code: 0 if all succeeded and at least one file processed.
    """

    _safe_mkdir(output_dir)

    excel_files: List[str] = []
    for name in os.listdir(input_dir):
        if _is_excel_file(name):
            excel_files.append(os.path.join(input_dir, name))

    if not excel_files:
        msg = f"No Excel files found in: {input_dir}"
        print(msg)
        _notify_windows("No Excel files", msg)
        return 2

    ok: List[str] = []
    failed: List[str] = []

    for excel_path in excel_files:
        base = os.path.splitext(os.path.basename(excel_path))[0]
        json_path = os.path.join(output_dir, base + ".json")
        try:
            excel_to_json(excel_path, json_file_path=json_path)
            ok.append(os.path.basename(excel_path))
        except Exception as e:
            failed.append(f"{os.path.basename(excel_path)} -> {e}")

    lines: List[str] = []
    lines.append(f"文件目录: {input_dir}")
    lines.append(f"输出目录: {output_dir}")
    lines.append(f"总数量: {len(excel_files)}  成功数量: {len(ok)}  失败数量: {len(failed)}")
    if ok:
        lines.append("\n成功文件:")
        lines.extend([f"- {n}" for n in ok])
    if failed:
        lines.append("\n失败文件:")
        lines.extend([f"- {n}" for n in failed])

    summary = "\n".join(lines)
    print("\n" + summary)

    if failed:
        _notify_windows("Excel转JSON失败", summary)
        return 1

    _notify_windows("Excel转JSON成功", summary)
    return 0


if __name__ == "__main__":
    base_dir = _default_input_dir()
    out_dir = os.path.join(base_dir, "Config")
    raise SystemExit(convert_all_excels_in_dir(base_dir, out_dir))
