import json
import os
from typing import Any, Dict, List, Tuple

import openpyxl


TYPE_INT = "int"
TYPE_STRING = "string"
TYPE_FLOAT = "float"
TYPE_FLOAT_ARR = "floatArr"
TYPE_INT_ARR = "intArr"
TYPE_STRING_ARR = "stringArr"
TYPE_FLOAT_ARR2 = "floatArr2"
TYPE_INT_ARR2 = "intArr2"
TYPE_STRING_ARR2 = "stringArr2"


def _flatten_dict(d: Dict[str, Any], parent_key: str = "", sep: str = "_") -> Dict[str, Any]:
    """Flatten nested dict using the same key joining style as before."""

    flat: Dict[str, Any] = {}
    for key, value in d.items():
        new_key = f"{parent_key}{sep}{key}" if parent_key else key
        if isinstance(value, dict):
            flat.update(_flatten_dict(value, new_key, sep=sep))
        else:
            flat[new_key] = value
    return flat


def _value_kind(value: Any) -> str | None:
    if value is None:
        return None

    if isinstance(value, bool):
        # Unity config usually treats bool as int 0/1; if you need bool, add a new type.
        return TYPE_INT

    if isinstance(value, int):
        return TYPE_INT

    if isinstance(value, float):
        return TYPE_FLOAT

    if isinstance(value, str):
        return TYPE_STRING

    if isinstance(value, list):
        if len(value) == 0:
            return None

        if all(isinstance(x, list) for x in value):
            # 2D
            saw_string = False
            all_numbers: List[float] = []
            for row in value:
                if not isinstance(row, list) or len(row) == 0:
                    continue
                for x in row:
                    if isinstance(x, str):
                        saw_string = True
                    elif isinstance(x, bool):
                        all_numbers.append(float(int(x)))
                    elif isinstance(x, (int, float)):
                        all_numbers.append(float(x))
                    else:
                        saw_string = True

            if saw_string:
                return TYPE_STRING_ARR2
            if any((not float(n).is_integer()) for n in all_numbers):
                return TYPE_FLOAT_ARR2
            return TYPE_INT_ARR2

        # 1D
        saw_string_1d = False
        all_numbers_1d: List[float] = []
        for x in value:
            if isinstance(x, str):
                saw_string_1d = True
            elif isinstance(x, bool):
                all_numbers_1d.append(float(int(x)))
            elif isinstance(x, (int, float)):
                all_numbers_1d.append(float(x))
            else:
                saw_string_1d = True

        if saw_string_1d:
            return TYPE_STRING_ARR
        if any((not float(n).is_integer()) for n in all_numbers_1d):
            return TYPE_FLOAT_ARR
        return TYPE_INT_ARR

    return TYPE_STRING


def _merge_types(current: str | None, incoming: str | None) -> str | None:
    if current is None:
        return incoming
    if incoming is None:
        return current

    if current == incoming:
        return current

    numeric = {TYPE_INT, TYPE_FLOAT}
    if current in numeric and incoming in numeric:
        return TYPE_FLOAT

    # array numeric widening
    if current == TYPE_INT_ARR and incoming == TYPE_FLOAT_ARR:
        return TYPE_FLOAT_ARR
    if current == TYPE_FLOAT_ARR and incoming == TYPE_INT_ARR:
        return TYPE_FLOAT_ARR
    if current == TYPE_INT_ARR2 and incoming == TYPE_FLOAT_ARR2:
        return TYPE_FLOAT_ARR2
    if current == TYPE_FLOAT_ARR2 and incoming == TYPE_INT_ARR2:
        return TYPE_FLOAT_ARR2

    # array string types
    if current == TYPE_STRING_ARR and incoming == TYPE_STRING_ARR:
        return TYPE_STRING_ARR
    if current == TYPE_STRING_ARR2 and incoming == TYPE_STRING_ARR2:
        return TYPE_STRING_ARR2

    return TYPE_STRING


def _infer_column_types(rows: List[Dict[str, Any]], keys: List[str]) -> Dict[str, str]:
    inferred: Dict[str, str | None] = {k: None for k in keys}

    for row in rows:
        for k in keys:
            kind = _value_kind(row.get(k))
            inferred[k] = _merge_types(inferred[k], kind)

    # default any unknowns to string
    return {k: (inferred[k] or TYPE_STRING) for k in keys}


def _to_excel_cell(value: Any, kind: str) -> Any:
    # IMPORTANT: keep empty string distinguishable from None
    if value is None:
        return None

    if kind == TYPE_STRING:
        # Excel tends to lose empty strings on reload (becomes blank/None).
        # Use a visible token to roundtrip empty string.
        if value == "":
            return '""'
        return str(value)

    if kind == TYPE_INT:
        if isinstance(value, bool):
            return int(value)
        if isinstance(value, (int, float)):
            return int(value)
        if isinstance(value, str) and value.strip() != "":
            return int(float(value))
        return None

    if kind == TYPE_FLOAT:
        if isinstance(value, (int, float)):
            return float(value)
        if isinstance(value, str) and value.strip() != "":
            return float(value)
        return None

    if kind in (TYPE_FLOAT_ARR, TYPE_INT_ARR, TYPE_STRING_ARR, TYPE_FLOAT_ARR2, TYPE_INT_ARR2, TYPE_STRING_ARR2):
        # Store arrays as JSON text for easy editing and stable parsing
        return json.dumps(value, ensure_ascii=False)

    return str(value)


def json_to_excel(json_path: str, excel_path: str = "game_config.xlsx") -> str:
    """JSON -> Excel with 3 header rows:

    Row 1: Chinese comments (left blank for you)
    Row 2: keys
    Row 3: types (int, string, float, floatArr, intArr, floatArr2, intArr2)
    Row 4+: data
    """

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    def _write(path: str) -> None:
        wb = openpyxl.Workbook()
        # remove default sheet
        wb.remove(wb.active)

        for top_key, top_value in data.items():
            if not (isinstance(top_value, list) and len(top_value) > 0):
                continue

            ws = wb.create_sheet(title=str(top_key)[:31])

            # flatten each row (nested dict -> key_key)
            flat_rows: List[Dict[str, Any]] = []
            keys: List[str] = []
            seen = set()
            for item in top_value:
                if not isinstance(item, dict):
                    continue
                flat = _flatten_dict(item)
                flat_rows.append(flat)
                for k in flat.keys():
                    if k not in seen:
                        seen.add(k)
                        keys.append(k)

            col_types = _infer_column_types(flat_rows, keys)

            # row 1: comment placeholders
            for col, _k in enumerate(keys, start=1):
                ws.cell(row=1, column=col).value = ""

            # row 2: keys
            for col, k in enumerate(keys, start=1):
                ws.cell(row=2, column=col).value = k

            # row 3: types
            for col, k in enumerate(keys, start=1):
                ws.cell(row=3, column=col).value = col_types[k]

            # data rows
            for r_i, row in enumerate(flat_rows, start=4):
                for c_i, k in enumerate(keys, start=1):
                    kind = col_types[k]
                    ws.cell(row=r_i, column=c_i).value = _to_excel_cell(row.get(k), kind)

            # freeze the three header rows
            ws.freeze_panes = "A4"

        wb.save(path)

    try:
        _write(excel_path)
        print(f"JSON转Excel成功！输出文件：{excel_path}")
        return excel_path
    except PermissionError:
        base, ext = os.path.splitext(excel_path)
        alt_path = f"{base}_new{ext or '.xlsx'}"
        _write(alt_path)
        print(f"JSON转Excel成功！输出文件：{alt_path}")
        print(f"提示：原文件可能被占用，已改写入：{alt_path}")
        return alt_path


if __name__ == "__main__":
    JSON_INPUT_PATH = "game_config.json"
    EXCEL_OUTPUT_PATH = "game_config.xlsx"
    json_to_excel(JSON_INPUT_PATH, EXCEL_OUTPUT_PATH)
